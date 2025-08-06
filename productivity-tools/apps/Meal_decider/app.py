from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os

# File paths
MENU_DB_FILE = "menu_database.xlsx"
GROCERY_DB_FILE = "grocery_database.xlsx"

app = Flask(__name__)

# Initialize the databases
def initialize_databases():
    if not os.path.exists(MENU_DB_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Menu"
        ws.append(["Dish", "Ingredients", "Time (mins)", "Category"])  # Headers
        wb.save(MENU_DB_FILE)

    if not os.path.exists(GROCERY_DB_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Groceries"
        ws.append(["Ingredient", "Quantity"])  # Headers
        wb.save(GROCERY_DB_FILE)

# Helper functions
def get_menu():
    wb = openpyxl.load_workbook(MENU_DB_FILE)
    ws = wb["Menu"]
    menu = [{"dish": row[0], "ingredients": row[1], "time": row[2], "category": row[3]} for row in ws.iter_rows(min_row=2, values_only=True)]
    return menu

def get_groceries():
    wb = openpyxl.load_workbook(GROCERY_DB_FILE)
    ws = wb["Groceries"]
    groceries = [{"ingredient": row[0], "quantity": row[1]} for row in ws.iter_rows(min_row=2, values_only=True)]
    return groceries

def add_to_menu(dish, ingredients, time, category):
    wb = openpyxl.load_workbook(MENU_DB_FILE)
    ws = wb["Menu"]
    ws.append([dish, ingredients, time, category])
    wb.save(MENU_DB_FILE)

def add_to_grocery_list(ingredient, quantity):
    wb = openpyxl.load_workbook(GROCERY_DB_FILE)
    ws = wb["Groceries"]
    ws.append([ingredient, quantity])
    wb.save(GROCERY_DB_FILE)

def choose_dishes(time_available):
    groceries = {item["ingredient"].strip().lower(): item["quantity"] for item in get_groceries()}
    menu = get_menu()
    full_matches = []
    partial_matches = []

    for item in menu:
        if int(item["time"]) > int(time_available):
            continue
        ingredient_list = [ingredient.strip().lower() for ingredient in item["ingredients"].split(",")]
        matched_ingredients = [ingredient for ingredient in ingredient_list if ingredient in groceries]

        if len(matched_ingredients) == len(ingredient_list):
            # Full match: All ingredients are available
            full_matches.append(item)
        elif matched_ingredients:
            # Partial match: Some ingredients are available
            item["matched_count"] = len(matched_ingredients)
            item["total_count"] = len(ingredient_list)
            partial_matches.append(item)

    # Sort partial matches by the number of matched ingredients (descending)
    partial_matches.sort(key=lambda x: x["matched_count"], reverse=True)

    return full_matches, partial_matches

# Routes
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/menu")
def menu():
    menu = get_menu()
    return render_template("menu.html", menu=menu)

@app.route("/add_menu", methods=["POST"])
def add_menu():
    dish = request.form["dish"]
    ingredients = request.form["ingredients"]
    time = request.form["time"]
    category = request.form["category"]
    add_to_menu(dish, ingredients, time, category)
    return redirect(url_for("menu"))

@app.route("/groceries")
def groceries():
    groceries = get_groceries()
    return render_template("groceries.html", groceries=groceries)

@app.route("/add_grocery", methods=["POST"])
def add_grocery():
    ingredient = request.form["ingredient"]
    quantity = request.form["quantity"]
    add_to_grocery_list(ingredient, quantity)
    return redirect(url_for("groceries"))

@app.route("/choose", methods=["GET", "POST"])
def choose():
    if request.method == "POST":
        time_available = request.form["time"]
        dishes = choose_dishes(time_available)
        return render_template("choose.html", dishes=dishes, time_available=time_available)
    return render_template("choose.html", dishes=None)

if __name__ == "__main__":
    initialize_databases()
    app.run(debug=True, host="0.0.0.0", port=5001)