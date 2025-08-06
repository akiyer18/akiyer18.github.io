import openpyxl
import pandas as pd
import random
import os

# File paths
MENU_DB_FILE = "menu_database.xlsx"
GROCERY_DB_FILE = "grocery_database.xlsx"

# Initialize the databases
def initialize_databases():
    # Initialize menu database
    if not os.path.exists(MENU_DB_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Menu"
        ws.append(["Dish", "Ingredients", "Time (mins)", "Category"])  # Headers
        wb.save(MENU_DB_FILE)

    # Initialize grocery database
    if not os.path.exists(GROCERY_DB_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Groceries"
        ws.append(["Ingredient", "Quantity"])  # Headers
        wb.save(GROCERY_DB_FILE)

def add_to_grocery_list():
    ingredient = input("Enter the ingredient name: ")
    quantity = input("Enter the quantity (e.g., 1kg, 500g, 2pcs): ")

    wb = openpyxl.load_workbook(GROCERY_DB_FILE)
    ws = wb["Groceries"]
    ws.append([ingredient, quantity])
    wb.save(GROCERY_DB_FILE)
    print(f"{ingredient} has been added to the grocery list!")

def check_groceries():
    wb = openpyxl.load_workbook(GROCERY_DB_FILE)
    ws = wb["Groceries"]
    groceries = {row[0].strip().lower(): row[1] for row in ws.iter_rows(min_row=2, values_only=True)}
    return groceries

def choose_what_to_eat():
    time_available = input("Enter the time you have (in minutes): ")
    groceries = check_groceries()

    wb = openpyxl.load_workbook(MENU_DB_FILE)
    ws = wb["Menu"]
    possible_dishes = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        dish, ingredients, time, category = row
        if int(time) > int(time_available):
            continue
        ingredient_list = [ingredient.strip().lower() for ingredient in ingredients.split(",")]
        if all(ingredient in groceries for ingredient in ingredient_list):
            possible_dishes.append((dish, time))

    if possible_dishes:
        print("\nYou can prepare the following dishes:")
        for dish, time in possible_dishes:
            print(f"- {dish} (Time: {time} mins)")
    else:
        print("\nInsufficient groceries to prepare any dish.")
        print("Options:")
        print("1. Add items to the grocery list")
        print("2. Exit (Go buy more groceries!)")
        choice = input("Enter your choice: ")
        if choice == "1":
            add_to_grocery_list()
        elif choice == "2":
            print("Goodbye! Time to restock your kitchen!")

def add_to_menu():
    dish = input("Enter the dish name: ")
    ingredients = input("Enter the ingredients (comma-separated): ")
    time = input("Enter the time required to prepare (in minutes): ")
    category = input("Enter the category (e.g., Breakfast, Lunch, Dinner, Snack): ")

    wb = openpyxl.load_workbook(MENU_DB_FILE)
    ws = wb["Menu"]
    ws.append([dish, ingredients, time, category])
    wb.save(MENU_DB_FILE)
    print(f"{dish} has been added to the menu!")

def main():
    initialize_databases()
    while True:
        print("\nMenu Options:")
        print("1. Add a dish to the menu")
        print("2. Add items to the grocery list")
        print("3. Choose what to eat")
        print("4. Exit")
        choice = input("Enter your choice: ")

        if choice == "1":
            add_to_menu()
        elif choice == "2":
            add_to_grocery_list()
        elif choice == "3":
            choose_what_to_eat()
        elif choice == "4":
            print("Goodbye! Happy cooking!")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()